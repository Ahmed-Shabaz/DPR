import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import ttkbootstrap as tb
import sqlite3
import os
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
from ttkbootstrap.widgets import DateEntry
from tkinter import font
from PIL import Image, ImageTk

class AdminLoginDialog(simpledialog.Dialog):
    def body(self, master):
        self.title("Admin Login")

        # Set the icon for the dialog
        icon_path = os.path.join(os.path.dirname(__file__), 'DPR_5.ico')
        if os.path.exists(icon_path):
            self.wm_iconbitmap(icon_path)
            
        self.username_entry = ttk.Entry(master, show='*')  # initially masked
            
        ttk.Label(master, text="Enter Admin Username:").grid(row=0, column=0, sticky='w', pady=(10, 2))
        self.username_entry.grid(row=1, column=0, padx=10, pady=(0, 10))
        self.eye_user_btn = ttk.Button(master, text='👁', width=3, command=self.toggle_username)
        self.eye_user_btn.grid(row=1, column=1, padx=(0, 10), pady=(0, 10))
        ttk.Label(master, text="Enter Admin Password:").grid(row=2, column=0, sticky='w', pady=(10, 2))
        self.password_entry = ttk.Entry(master, show='*')
        self.password_entry.grid(row=3, column=0, padx=10, pady=(0, 10))
        self.eye_pass_btn = ttk.Button(master, text='👁', width=3, command=self.toggle_password)
        self.eye_pass_btn.grid(row=3, column=1, padx=(0, 10), pady=(0, 10))
        return self.username_entry  # initial focus
    
    def toggle_username(self):
        if self.username_entry.cget('show') == '':
            self.username_entry.config(show='*')
            self.eye_user_btn.config(text='👁')
        else:
            self.username_entry.config(show='')
            self.eye_user_btn.config(text='🚫')
    
    def toggle_password(self):
        if self.password_entry.cget('show') == '':
            self.password_entry.config(show='*')
            self.eye_pass_btn.config(text='👁')
        else:
            self.password_entry.config(show='')
            self.eye_pass_btn.config(text='🚫')
    
    def apply(self):
        self.result = (self.username_entry.get(), self.password_entry.get())


class ManagerLoginDialog(simpledialog.Dialog):
    def body(self, master):
        # Set the icon for the dialog
        icon_path = os.path.join(os.path.dirname(__file__), 'DPR_5.ico')
        if os.path.exists(icon_path):
            self.wm_iconbitmap(icon_path)
            
        self.title("Manager Login")
        ttk.Label(master, text="Enter Manager Username:").grid(row=0, column=0, sticky='w', pady=(10, 2))
        self.username_entry = ttk.Entry(master, show='*')  # initially masked
        self.username_entry.grid(row=1, column=0, padx=10, pady=(0, 10))
        self.eye_user_btn = ttk.Button(master, text='👁', width=3, command=self.toggle_username)
        self.eye_user_btn.grid(row=1, column=1, padx=(0, 10), pady=(0, 10))
        ttk.Label(master, text="Enter Manager Password:").grid(row=2, column=0, sticky='w', pady=(10, 2))
        self.password_entry = ttk.Entry(master, show='*')
        self.password_entry.grid(row=3, column=0, padx=10, pady=(0, 10))
        self.eye_pass_btn = ttk.Button(master, text='👁', width=3, command=self.toggle_password)
        self.eye_pass_btn.grid(row=3, column=1, padx=(0, 10), pady=(0, 10))
        return self.username_entry  # initial focus
    
    def toggle_username(self):
        if self.username_entry.cget('show') == '':
            self.username_entry.config(show='*')
            self.eye_user_btn.config(text='👁')
        else:
            self.username_entry.config(show='')
            self.eye_user_btn.config(text='🚫')
    
    def toggle_password(self):
        if self.password_entry.cget('show') == '':
            self.password_entry.config(show='*')
            self.eye_pass_btn.config(text='👁')
        else:
            self.password_entry.config(show='')
            self.eye_pass_btn.config(text='🚫')
    
    def apply(self):
        self.result = (self.username_entry.get(), self.password_entry.get())

class MonthYearDialog(simpledialog.Dialog):

    def body(self, master):

        self.title("Select Month and Year")
        ttk.Label(master, text="Select Month:").grid(row=0, column=0, sticky='w', pady=(10, 2))
        self.month_cb = ttk.Combobox(master, values=[f"{i:02d}" for i in range(1, 13)], state='readonly', width=10)
        self.month_cb.grid(row=1, column=0, padx=10, pady=(0, 10))
        self.month_cb.set(datetime.now().strftime("%m"))

        ttk.Label(master, text="Enter Year (YYYY):").grid(row=2, column=0, sticky='w', pady=(10, 2))
        self.year_entry = ttk.Entry(master, width=13)
        self.year_entry.grid(row=3, column=0, padx=10, pady=(0, 10))
        self.year_entry.insert(0, datetime.now().strftime("%Y"))

        return self.month_cb

    def validate(self):
        month = self.month_cb.get()
        year = self.year_entry.get()
        if not (month and year.isdigit() and len(year) == 4):
            messagebox.showwarning("Invalid Input", "Please enter valid Month and 4-digit Year.")
            return False
        return True

    def apply(self):
        self.result = (self.month_cb.get(), self.year_entry.get())

class DPRApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BMS DPR")
        self.style = tb.Style('solar')

        icon_path = os.path.join(os.path.dirname(__file__), 'DPR_5.ico')
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)
        
        self.style.configure("Bold.TButton", font=("Segoe UI", 10, "bold"))
        self.style.configure('NoBorder.TButton', borderwidth=0, focusthickness=0, focuscolor='')

        def create_style(style_name, bg_color):
            self.style.configure(style_name, background=bg_color, foreground='white', font=('Segoe UI', 10, 'bold'), borderwidth=0)
            self.style.map(style_name, background=[('active', bg_color)], foreground=[('active', 'white')])

        create_style('Admin.TButton', '#c28330')
        create_style('Manager.TButton', '#c28330')
        create_style('Submit.TButton', '#0b97ba')
        create_style('DeleteNav.TButton', '#2a6b87')
        create_style('BackHome.TButton', '#3d5e87')
        create_style('DeleteBack.TButton', '#c28330')
        create_style('DeleteRecords.TButton', '#9a709c')
        create_style('SubmitForm.TButton', '#007acc')
        create_style('BackForm.TButton', '#f59627')
        create_style('Export.TButton', '#28a745')

        # Base path for files
        self.admin_base_path = r"C:\Users\shabaz\Desktop\DPR"
        self.manager_base_path = r"C:\Users\shabaz\Desktop\DPR"

        self.role = None  # Admin or Manager
        self.base_path = None
        self.db_path = None
        self.input_excel_path = None
        self.login_details_path = None

        self.selected_date = None
        self.emp_codes = {}
        self.in_time = None

        # Initialize in_time_label and in_time_value_label to None to avoid attribute errors
        self.in_time_label = None
        self.in_time_value_label = None

        # Load logos for roles
        self.load_role_logos()

        self.show_home_page()

    def load_role_logos(self):
        try:
            base_dir = os.path.dirname(__file__)
            admin_logo_path = os.path.join(base_dir, 'admin_logo.png')
            manager_logo_path = os.path.join(base_dir, 'admin_logo.png')

            if os.path.exists(admin_logo_path):
                self.admin_logo_img = tk.PhotoImage(file=admin_logo_path)
            else:
                self.admin_logo_img = None

            if os.path.exists(manager_logo_path):
                self.manager_logo_img = tk.PhotoImage(file=manager_logo_path)
            else:
                self.manager_logo_img = None
        except Exception:
            self.admin_logo_img = None
            self.manager_logo_img = None

    def show_home_page(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        home_frame = ttk.Frame(self.root)
        home_frame.pack(padx=20, pady=40)

        base_dir = os.path.dirname(__file__)
        logo_image_path = os.path.join(base_dir, 'BMS_Logo.png')
        logo_image = Image.open(logo_image_path)
        logo_image = logo_image.resize((120, 40), Image.Resampling.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)

        role_frame = tk.Frame(home_frame)
        role_frame.pack(pady=20)

        logo_label = tk.Label(role_frame, image=logo_photo)
        logo_label.image = logo_photo
        logo_label.pack(side="left", padx=(0, 10))

        ttk.Label(home_frame, text="Select Role", font=("Segoe UI", 24, "bold")).pack(pady=20)

        roles_frame = ttk.Frame(home_frame)
        roles_frame.pack()

        admin_frame = ttk.Frame(roles_frame)
        admin_frame.grid(row=0, column=0, padx=40, pady=10)

        if self.admin_logo_img:
            admin_logo_label = ttk.Label(admin_frame, image=self.admin_logo_img)
            admin_logo_label.pack(pady=(0, 10))
        else:
            admin_logo_label = ttk.Label(admin_frame, text="Admin Logo", font=("Segoe UI", 14))
            admin_logo_label.pack(pady=(0, 10))

        self.style.configure("Bold.TButton", font=("Segoe UI", 10, "bold"))
        admin_button = ttk.Button(admin_frame, text="Admin", command=self.admin_login_prompt, bootstyle='primary', style="Bold.TButton", width=20)
        admin_button.pack()

        manager_frame = ttk.Frame(roles_frame)
        manager_frame.grid(row=0, column=1, padx=40, pady=10)

        if self.manager_logo_img:
            manager_logo_label = ttk.Label(manager_frame, image=self.manager_logo_img)
            manager_logo_label.pack(pady=(0, 10))
        else:
            manager_logo_label = ttk.Label(manager_frame, text="Manager Logo", font=("Segoe UI", 14))
            manager_logo_label.pack(pady=(0, 10))

        self.style.configure("Bold.TButton", font=("Segoe UI", 10, "bold"))
        manager_button = ttk.Button(manager_frame, text="Manager", command=self.manager_login_prompt, bootstyle='primary', style="Bold.TButton", width=20)
        manager_button.pack()

    def admin_login_prompt(self):
        dlg = AdminLoginDialog(self.root)
        if dlg.result is None:
            return
        username, password = dlg.result
        if username == "adm" and password == "adm123":
            self.admin_init()
        else:
            messagebox.showerror("Incorrect Credentials", "Incorrect Username or Password")

    def manager_login_prompt(self):
        dlg = ManagerLoginDialog(self.root)
        if dlg.result is None:
            return
        username, password = dlg.result
        if username == "man" and password == "man123":
            self.manager_init()
        else:
            messagebox.showerror("Incorrect Credentials", "Incorrect Username or Password")

    def admin_init(self):
        self.role = "Admin"
        self.base_path = self.admin_base_path
        self.db_path = os.path.join(self.base_path, "DPR.db")
        self.input_excel_path = os.path.join(self.base_path, "Projects_&_Users_Input.xlsx")
        self.login_details_path = os.path.join(self.base_path, "Login Details.xlsx")
        os.makedirs(self.base_path, exist_ok=True)

        self.load_excel_data()
        self.create_database()
        self.show_calendar()

    def manager_init(self):
        self.role = "Manager"
        self.base_path = self.manager_base_path
        self.db_path = os.path.join(self.base_path, "DPR.db")
        self.input_excel_path = os.path.join(self.base_path, "Projects_&_Users_Input.xlsx")
        self.login_details_path = os.path.join(self.base_path, "Login Details.xlsx")
        os.makedirs(self.base_path, exist_ok=True)

        self.load_excel_data()
        self.create_database()
        self.selected_date = datetime.now().strftime('%d-%m-%Y')
        self.create_widgets()

    def show_calendar(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        self.calendar_frame = ttk.Frame(self.root)
        self.calendar_frame.pack(padx=30, pady=50)

        ttk.Label(self.calendar_frame,
                  text="Select a Date",
                  font=("Segoe UI", 18, "bold")).pack(pady=(60, 20))

        self.date_entry = DateEntry(self.calendar_frame,
                                    width=30,
                                    bootstyle='info',
                                    firstweekday=6,
                                    dateformat="%d-%m-%Y")
        self.date_entry.pack(pady=12)

        self.date_submit_btn = ttk.Button(self.calendar_frame, text="Submit", command=self.submit_date, style='Submit.TButton')
        self.date_submit_btn.pack(pady=10)

        self.goto_delete_btn = ttk.Button(self.calendar_frame, text="Navigate to Delete Record", command=self.show_delete_record_page, style='DeleteNav.TButton')
        self.goto_delete_btn.pack(pady=10)

        self.back_home_btn = ttk.Button(self.calendar_frame, text="← Back to Home Page", command=self.show_home_page, style='BackHome.TButton')
        self.back_home_btn.pack(pady=14)

    def submit_date(self):
        self.selected_date = self.date_entry.entry.get()
        if not self.selected_date:
            messagebox.showwarning("Input Error", "Please select a date.")
            return
        try:
            selected_date_obj = datetime.strptime(self.selected_date, "%d-%m-%Y")
        except Exception:
            messagebox.showwarning("Input Error", "Invalid date format selected.")
            return
        today_date_obj = datetime.now()

        if selected_date_obj > today_date_obj:
            messagebox.showwarning("Input Error", "You cannot select a future date.")
            return

        self.calendar_frame.destroy()
        self.create_widgets()

    def show_delete_record_page(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        delete_frame = ttk.Frame(self.root)
        delete_frame.pack(fill='both', expand=True, padx=20, pady=20)

        ttk.Label(delete_frame, text="To Delete Records", font=("Segoe UI", 20, "bold")).pack(pady=30)

        ttk.Label(delete_frame, text="From", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.from_date_entry = DateEntry(delete_frame,
                                        width=30,
                                        bootstyle='info',
                                        firstweekday=6,
                                        dateformat="%d-%m-%Y")
        self.from_date_entry.pack(pady=10)

        ttk.Label(delete_frame, text="To", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.to_date_entry = DateEntry(delete_frame,
                                      width=30,
                                      bootstyle='info',
                                      firstweekday=6,
                                      dateformat="%d-%m-%Y")
        self.to_date_entry.pack(pady=10)

        ttk.Label(delete_frame, text="Project Name", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.delete_project_name_cb = ttk.Combobox(delete_frame, values=self.all_project_names, state='readonly', width=30)
        self.delete_project_name_cb.pack(pady=10)
        self.delete_project_name_cb.set('Select Project')

        ttk.Label(delete_frame, text="Team Member (optional)", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.delete_team_member_cb = ttk.Combobox(delete_frame, values=list(self.emp_codes.keys()), state='readonly', width=30)
        self.delete_team_member_cb.pack(pady=10)
        self.delete_team_member_cb.set('Select Team Member')

        self.delete_records_button = ttk.Button(delete_frame, text="Delete Records", command=self.delete_records, style='DeleteRecords.TButton')
        self.delete_records_button.pack(pady=10)

        back_btn = ttk.Button(delete_frame, text="← Back", command=self.back_from_delete_record, style='DeleteBack.TButton')
        back_btn.pack(pady=10)

    def back_from_delete_record(self):
        if self.role == "Admin":
            self.show_calendar()
        else:
            self.show_home_page()

    def delete_records(self):
        from_date = self.from_date_entry.entry.get()
        to_date = self.to_date_entry.entry.get()
        project_name = self.delete_project_name_cb.get()
        team_member = self.delete_team_member_cb.get()

        if not from_date or not to_date or project_name == 'Select Project':
            messagebox.showwarning("Input Error", "Please select From Date, To Date, and Project Name.")
            return

        try:
            from_date_obj = datetime.strptime(from_date, "%d-%m-%Y")
            to_date_obj = datetime.strptime(to_date, "%d-%m-%Y")
        except Exception:
            messagebox.showwarning("Input Error", "Invalid date format selected.")
            return

        if from_date_obj > to_date_obj:
            messagebox.showwarning("Input Error", "From Date must be earlier than To Date.")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Adjust the SQL query to delete based on project name and optionally team member
        if team_member == 'Select Team Member':
            cursor.execute('''DELETE FROM DPR 
                              WHERE Date BETWEEN ? AND ? AND Project_Name = ?''',
                           (from_date, to_date, project_name))
        else:
            cursor.execute('''DELETE FROM DPR 
                              WHERE Date BETWEEN ? AND ? AND Project_Name = ? AND "Team Members" = ?''',
                           (from_date, to_date, project_name, team_member))

        conn.commit()
        deleted_count = cursor.rowcount
        conn.close()

        if deleted_count > 0:
            messagebox.showinfo("Success", f"Deleted {deleted_count} records successfully!")
        else:
            messagebox.showinfo("No Records Deleted", "No records found for the selected criteria.")

    def load_excel_data(self):
        try:
            workbook = load_workbook(self.input_excel_path)

            user_list_sheet = workbook['User_List']
            project_name_sheet = workbook['Project_Name']
            process_type_sheet = workbook['Process_Type']

            self.leads = sorted(set(
                user_list_sheet.cell(row=row, column=1).value
                for row in range(2, user_list_sheet.max_row + 1)
                if user_list_sheet.cell(row=row, column=1).value is not None
            ))

            self.user_names_for_projects = sorted(set(
                project_name_sheet.cell(row=row, column=1).value
                for row in range(2, project_name_sheet.max_row + 1)
                if project_name_sheet.cell(row=row, column=1).value is not None
            ))

            self.process_types = sorted(set(
                process_type_sheet.cell(row=row, column=1).value
                for row in range(2, process_type_sheet.max_row + 1)
                if process_type_sheet.cell(row=row, column=1).value is not None
            ))

            self.all_project_names = sorted(set(
                project_name_sheet.cell(row=row, column=3).value
                for row in range(2, project_name_sheet.max_row + 1)
                if project_name_sheet.cell(row=row, column=3).value is not None
            ))

            self.emp_codes = {}
            for row in range(2, user_list_sheet.max_row + 1):
                name = user_list_sheet.cell(row=row, column=2).value
                emp_code = user_list_sheet.cell(row=row, column=3).value
                if name and emp_code:
                    self.emp_codes[name] = emp_code

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel data:\n{e}")
            self.leads = []
            self.process_types = []
            self.all_project_names = []

    def create_database(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS DPR (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Date TEXT,
            Lead TEXT,
            "Team Members" TEXT,
            Project_Name TEXT,
            Process_Type TEXT,
            Other_process_details TEXT,
            Count INTEGER,
            Attendance TEXT,
            Leave_Type TEXT,
            Half_Day_Status TEXT,
            Permission_Details TEXT,
            Comments TEXT,
            In_Time TEXT,
            Entered_By TEXT
        )''')
        conn.commit()
        cursor.execute("PRAGMA table_info(DPR)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'Entered_By' not in columns:
            cursor.execute("ALTER TABLE DPR ADD COLUMN Entered_By TEXT")
            conn.commit()
        conn.close()

    def create_widgets(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill='x', pady=(10, 0), padx=10)

        ttk.Label(top_frame, text="Daily Production Report", font=("Segoe UI", 20, "bold")).pack(anchor='center', pady=(80, 0))

        ttk.Label(top_frame, text=f"Date: {self.selected_date}", font=("Segoe UI", 12)).pack(side='right', padx=(0, 40))

        self.form_frame = ttk.Frame(self.root)
        self.form_frame.pack(padx=20, pady=15)

        self.form_frame = tk.Frame(self.form_frame, borderwidth=10, relief="groove", padx=20, pady=20)
        self.form_frame.pack(fill='both', expand=True)

        label_font = font.Font(family="Helvetica", size=10, weight="bold")

        # LEAD
        ttk.Label(self.form_frame, text="Lead*", font=label_font).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.group_lead_cb = ttk.Combobox(self.form_frame, values=self.leads, state='readonly', width=30)
        self.group_lead_cb.grid(row=1, column=0, padx=5)
        self.group_lead_cb.set('Select Lead')
        self.group_lead_cb.bind("<<ComboboxSelected>>", self.update_user_names)

        # TEAM MEMBERS
        ttk.Label(self.form_frame, text="Team Members*", font=label_font).grid(row=0, column=1, sticky='w', padx=5, pady=5)
        self.user_name_cb = ttk.Combobox(self.form_frame, values=[], state='readonly', width=30)
        self.user_name_cb.grid(row=1, column=1, padx=5)
        self.user_name_cb.set('Select Team Member')
        self.user_name_cb.bind("<<ComboboxSelected>>", self.on_user_name_selected)

        # PROJECT NAME
        ttk.Label(self.form_frame, text="Project Name*", font=label_font).grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.project_name_cb = ttk.Combobox(self.form_frame, values=self.all_project_names, state='readonly', width=30)
        self.project_name_cb.grid(row=1, column=2, padx=5)
        self.project_name_cb.set('Select Project')

        # PROCESS TYPE
        ttk.Label(self.form_frame, text="Process Type*", font=label_font).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.process_type_cb = ttk.Combobox(self.form_frame, values=self.process_types, state='readonly', width=30)
        self.process_type_cb.grid(row=3, column=0, padx=5)
        self.process_type_cb.set('Select Process')  # Initially empty
        self.process_type_cb.bind("<<ComboboxSelected>>", self.toggle_process_details)


        # COUNT
        ttk.Label(self.form_frame, text="Count*", font=label_font).grid(row=2, column=2, sticky='w', padx=5, pady=5)
        self.count_entry = ttk.Entry(self.form_frame, width=33)
        self.count_entry.grid(row=3, column=2, padx=5)

        # ATTENDANCE
        ttk.Label(self.form_frame, text="Attendance*", font=label_font).grid(row=2, column=1, sticky='w', padx=5, pady=5)
        self.attendance_cb = ttk.Combobox(self.form_frame, values=["Present", "On Leave", "Permission", "Half Day"], state='readonly', width=30)
        self.attendance_cb.grid(row=3, column=1, padx=5)
        self.attendance_cb.set("Select Attendance")
        self.attendance_cb.bind("<<ComboboxSelected>>", self.toggle_leave_options)

        # CONDITIONAL FIELDS
        self.leave_type_label = ttk.Label(self.form_frame, text="Leave Type*", font=label_font)
        self.leave_type_combobox = ttk.Combobox(self.form_frame, values=["Planned Leave", "Unplanned Leave"], state='readonly', width=30)
        self.leave_type_combobox.set("Select Leave Type")

        self.type_label = ttk.Label(self.form_frame, text="Half Day Status*", font=label_font)
        self.type_combobox = ttk.Combobox(self.form_frame, values=["Planned", "Unplanned"], state='readonly', width=30)
        self.type_combobox.set("Select Half Day Status")

        self.permission_details_label = ttk.Label(self.form_frame, text="Permission Details*", font=label_font)
        self.permission_details_entry = ttk.Entry(self.form_frame, width=33)

        self.other_process_details_label = ttk.Label(self.form_frame, text="Other Process Details*", font=label_font)
        self.other_process_details_entry = ttk.Entry(self.form_frame, width=33)

        if self.role == "Manager":
            self.in_time_label = ttk.Label(self.form_frame, text="InTime:", font=label_font)
            # self.in_time_value_label = ttk.Label(self.form_frame, text="", font=label_font)
        else:
            self.in_time_label = None
            self.in_time_value_label = None

        ttk.Label(self.form_frame, text="Comments", font=label_font).grid(row=4, column=2, sticky='w', padx=5, pady=5)
        self.comments_entry = ttk.Entry(self.form_frame, width=33)
        self.comments_entry.grid(row=5, column=2, padx=5, pady=5)

        self.submit_button = ttk.Button(self.form_frame, text="Submit", width=20, command=self.submit_entry, style='SubmitForm.TButton')
        self.submit_button.grid(row=6, column=1, pady=20)

        button_frame = ttk.Frame(self.root)
        button_frame.pack(pady=(0, 15))

        if self.role == "Admin":
            back_command = self.show_calendar
        else:
            back_command = self.show_home_page

        self.back_button = ttk.Button(button_frame, text="← Back", command=back_command, style='BackForm.TButton')
        self.back_button.pack(side='left', padx=10)

        # Update: In Admin role, rename export button text to "Export All" instead of "Daily Export"
        if self.role == "Admin":
            export_button_text = "Export Data"
        else:
            export_button_text = "Daily Export"

        self.export_button = ttk.Button(button_frame, text=export_button_text, command=self.export_to_excel, style='Export.TButton')
        self.export_button.pack(side='left', padx=10)

        if self.role == "Admin":
            self.customize_button = ttk.Button(button_frame, text="Custom Export", command=self.customize_export, style='Export.TButton')
            self.customize_button.pack(side='left', padx=10)

        if self.role == "Manager":
            self.monthly_export_button = ttk.Button(button_frame, text="Monthly Export", command=self.export_monthly_data, style='Export.TButton')
            self.monthly_export_button.pack(side='left', padx=10)

            
    def toggle_process_details(self, event):
        self.other_process_details_label.grid_forget()
        self.other_process_details_entry.grid_forget()
        if self.process_type_cb.get() == "OTHERS":
            self.other_process_details_label.grid(row=4, column=0, sticky='w', padx=5, pady=5)
            self.other_process_details_entry.grid(row=5, column=0, padx=5)

    def update_user_names(self, event):
        selected_lead = self.group_lead_cb.get()
        if selected_lead != "Select Lead":
            try:
                workbook = load_workbook(self.input_excel_path)
                user_list_sheet = workbook['User_List']
                users = set()
                for row in range(2, user_list_sheet.max_row + 1):
                    lead_cell = user_list_sheet.cell(row=row, column=1).value
                    name_cell = user_list_sheet.cell(row=row, column=2).value
                    if lead_cell == selected_lead and name_cell is not None:
                        users.add(name_cell)
                users = sorted(users)
                self.user_name_cb['values'] = users
                if users:
                    self.user_name_cb.current(0)
                    self.on_user_name_selected(None)
                    self.attendance_cb.set("Present")
                else:
                    if self.in_time_label:
                        self.in_time_label.grid_forget()
                    if self.in_time_value_label:
                        self.in_time_value_label.grid_forget()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load user names:\n{e}")
                self.user_name_cb['values'] = []
                self.user_name_cb.set('Select Team Member')
                self.project_name_cb['values'] = []
                self.project_name_cb.set('Select Project')
                if self.in_time_label:
                    self.in_time_label.grid_forget()
                if self.in_time_value_label:
                    self.in_time_value_label.grid_forget()

        # Set Process Type to "PR" only after a Lead is selected
        self.process_type_cb.set('PR')
        self.toggle_process_details(None)



    def on_user_name_selected(self, event):
        selected_user = self.user_name_cb.get()
        if selected_user != "Select Team Member":
            try:
                self.get_in_time(selected_user)
                workbook = load_workbook(self.input_excel_path)
                project_name_sheet = workbook['Project_Name']
                user_projects = []
                for row in range(2, project_name_sheet.max_row + 1):
                    user_cell = project_name_sheet.cell(row=row, column=1).value
                    project_cell = project_name_sheet.cell(row=row, column=2).value
                    if user_cell == selected_user and project_cell is not None:
                        user_projects.append(project_cell)
                all_projects = list(self.all_project_names)
                remaining_projects = [p for p in all_projects if p not in user_projects]
                combined_projects = user_projects + remaining_projects
                self.project_name_cb['values'] = combined_projects
                if user_projects:
                    self.project_name_cb.set(user_projects[0])
                elif combined_projects:
                    self.project_name_cb.set(combined_projects[0])
                else:
                    self.project_name_cb.set('Select Project')
                if self.in_time_label:
                    self.in_time_label.grid(row=7, column=0, sticky='w', padx=5, pady=5)
                if self.in_time_value_label:
                    self.in_time_value_label.grid(row=7, column=1, sticky='w', padx=5, pady=5)
                    self.in_time_value_label.config(text=self.in_time)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load project names for user:\n{e}")
                self.project_name_cb['values'] = self.all_project_names
                self.project_name_cb.set('Select Project')
                self.attendance_cb.set("Present")
                if self.in_time_label:
                    self.in_time_label.grid_forget()
                if self.in_time_value_label:
                    self.in_time_value_label.grid_forget()

        # Reset Process Type to "PR" when a new user is selected
        self.process_type_cb.set('PR')
        self.toggle_process_details(None)


    def get_in_time(self, user_name):
        try:
            emp_code = self.emp_codes.get(user_name)
            if emp_code is None:
                self.in_time = "Missed Punch / Absent"
            else:
                login_workbook = load_workbook(self.login_details_path)
                login_sheet = login_workbook.active

                self.in_time = None
                for row in range(12, login_sheet.max_row + 1):
                    cell_value = login_sheet.cell(row=row, column=3).value
                    if cell_value is not None and str(cell_value).strip() == str(emp_code).strip():
                        in_time_cell = login_sheet.cell(row=row, column=8).value
                        if isinstance(in_time_cell, datetime):
                            self.in_time = in_time_cell.strftime("%m/%d/%Y %I:%M:%S %p")
                        else:
                            self.in_time = str(in_time_cell).strip() if in_time_cell else None
                        break
                if not self.in_time:
                    self.in_time = "Missed Punch / Absent"
            if self.in_time_label is not None:
                self.in_time_label.config(text=f"InTime: {self.in_time}")
                self.in_time_label.grid(row=7, column=0, sticky='w', padx=5, pady=5, columnspan=3)

            if self.role == "Manager" and self.in_time_label is not None and self.in_time_value_label is not None:
                self.in_time_label.grid(row=7, column=0, sticky='w', padx=5, pady=5)
                self.in_time_value_label.grid(row=7, column=1, sticky='w', padx=5, pady=5)
                self.in_time_value_label.config(text=self.in_time)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load In Time data:\n{e}")
            self.in_time = "Missed Punch / Absent"
            if self.in_time_label is not None:
                self.in_time_label.grid(row=7, column=0, sticky='w', padx=5, pady=5)
                self.in_time_label.config(text=f"InTime: {self.in_time}")
            if self.in_time_value_label is not None:
                self.in_time_value_label.grid(row=7, column=1, sticky='w', padx=5, pady=5)
                self.in_time_value_label.config(text=self.in_time)

    def toggle_leave_options(self, event):
        attendance = self.attendance_cb.get()
        self.leave_type_label.grid_forget()
        self.leave_type_combobox.grid_forget()
        self.permission_details_label.grid_forget()
        self.permission_details_entry.grid_forget()
        self.type_label.grid_forget()
        self.type_combobox.grid_forget()
        self.count_entry.config(state='normal')

        if attendance == "On Leave":
            self.leave_type_label.grid(row=4, column=1, sticky='w', padx=5, pady=5)
            self.leave_type_combobox.grid(row=5, column=1, padx=5)
            self.count_entry.delete(0, tk.END)
            self.count_entry.insert(0, "0")
            self.count_entry.config(state='disabled')
        else:
            self.count_entry.delete(0, tk.END)
            self.count_entry.config(state='normal')

        if attendance == "Permission":
            self.permission_details_label.grid(row=4, column=1, sticky='w', padx=5, pady=5)
            self.permission_details_entry.grid(row=5, column=1, padx=5)

        if attendance == "Half Day":
            self.type_label.grid(row=4, column=1, sticky='w', padx=5, pady=5)
            self.type_combobox.grid(row=5, column=1, padx=5)

    def submit_entry(self):
        mandatory_fields = {
            "Lead": self.group_lead_cb.get(),
            "Team Member": self.user_name_cb.get(),
            "Project": self.project_name_cb.get(),
            "Process": self.process_type_cb.get(),
            "Attendance": self.attendance_cb.get(),
        }

        for field, value in mandatory_fields.items():
            if value.startswith("Select"):
                messagebox.showwarning("Input Error", f"{field} is mandatory.")
                return

        if self.process_type_cb.get() == "OTHERS" and not self.other_process_details_entry.get().strip():
            messagebox.showwarning("Input Error", "Please fill in Other Process Details.")
            return

        if self.attendance_cb.get() == "On Leave" and self.leave_type_combobox.get() == "Select Leave Type":
            messagebox.showwarning("Input Error", "Please select a Leave Type.")
            return

        if self.attendance_cb.get() == "Half Day" and self.type_combobox.get() == "Select Half Day Status":
            messagebox.showwarning("Input Error", "Please select Half Day Status.")
            return

        if self.attendance_cb.get() == "Permission" and not self.permission_details_entry.get().strip():
            messagebox.showwarning("Input Error", "Please provide Permission Details.")
            return

        if not self.count_entry.get().strip():
            messagebox.showwarning("Input Error", "Count is mandatory.")
            return

        try:
            count = int(self.count_entry.get().strip())
        except ValueError:
            messagebox.showwarning("Input Error", "Count must be an integer.")
            return

        leave_type_val = self.leave_type_combobox.get() if self.attendance_cb.get() == "On Leave" else ""
        half_day_status_val = self.type_combobox.get() if self.attendance_cb.get() == "Half Day" else ""
        permission_details_val = self.permission_details_entry.get() if self.attendance_cb.get() == "Permission" else ""

        in_time_val = self.in_time if self.in_time is not None else "Missed Punch / Absent"

        team_member = self.user_name_cb.get()

        # Check for duplicate entries
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''SELECT COUNT(*) FROM DPR WHERE 
                          Date = ? AND 
                          Lead = ? AND 
                          "Team Members" = ? AND 
                          Project_Name = ? AND 
                          Process_Type = ?''', 
                       (self.selected_date, 
                        self.group_lead_cb.get(), 
                        team_member, 
                        self.project_name_cb.get(), 
                        self.process_type_cb.get()))
        
        duplicate_count = cursor.fetchone()[0]
        if duplicate_count > 0:
            messagebox.showwarning("Duplicate Entry", f"You Already Submitted For '{team_member}'")
            conn.close()
            return

        # Proceed with the insertion if no duplicates found
        data = (
            self.selected_date,
            self.group_lead_cb.get(),
            team_member,
            self.project_name_cb.get(),
            self.process_type_cb.get(),
            self.other_process_details_entry.get() if self.process_type_cb.get() == "OTHERS" else "",
            count,
            self.attendance_cb.get(),
            leave_type_val,
            half_day_status_val,
            permission_details_val,
            self.comments_entry.get(),
            in_time_val,
            self.role  # Store the role here
        )

        cursor.execute('''INSERT INTO DPR (
            Date, Lead, "Team Members", Project_Name, Process_Type,
            Other_process_details, Count, Attendance, Leave_Type, Half_Day_Status, Permission_Details, Comments, In_Time, Entered_By
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', data)
        conn.commit()
        conn.close()

        messagebox.showinfo("BMS DPR", "Entry submitted successfully!")
        self.reset_form()

    def reset_form(self):
        self.group_lead_cb.set('Select Lead')
        self.user_name_cb.set('Select Team Member')
        self.project_name_cb.set('Select Project')
        self.process_type_cb.set('Select Process')  # Reset to empty
        self.count_entry.config(state='normal')
        self.count_entry.delete(0, tk.END)
        self.attendance_cb.set('Select Attendance')
        self.leave_type_combobox.set("Select Leave Type")
        self.type_combobox.set("Select Half Day Status")
        self.permission_details_entry.delete(0, tk.END)
        self.other_process_details_entry.delete(0, tk.END)
        self.comments_entry.delete(0, tk.END)
        if self.in_time_label:
            self.in_time_label.grid_forget()
        if self.in_time_value_label:
            self.in_time_value_label.grid_forget()
            self.in_time_value_label.config(text="")
        self.in_time = None
        self.toggle_leave_options(None)
        self.toggle_process_details(None)



    def export_to_excel(self):
        # For Admin role, export all data, for others export selected date data
        if self.role == "Admin":
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Normalize casing and trim spaces for Entered_By check for Admin records
            try:
                cursor.execute("UPDATE DPR SET In_Time = '' WHERE TRIM(UPPER(Entered_By)) = 'ADMIN'")
                conn.commit()
            except Exception as e:
                messagebox.showerror("Database Error", f"Failed to update In Time for Admin records:\n{e}")
                conn.close()
                return

            # Query all rows from the DPR table
            try:
                cursor.execute("""
                    SELECT 
                        Lead, "Team Members", Project_Name, Process_Type, Other_process_details,
                        Attendance, Leave_Type, Half_Day_Status, Permission_Details, Count, Comments, Date, In_Time, Entered_By
                    FROM DPR
                """)
            except Exception as e:
                messagebox.showerror("Database Error", f"Failed to fetch data for export:\n{e}")
                conn.close()
                return

            headers = [
                "SL.No", "Lead", "Team Members", "Project Name", "Process Type", "Other Process Details",
                "Attendance", "Leave Type", "Half Day Status", "Permission Details", "Count", "Comments", "Date", "In Time"
            ]

            folder_selected = filedialog.askdirectory(title="Select Folder to Save Excel File")
            if not folder_selected:
                messagebox.showwarning("Export Cancelled", "No folder was selected. Export cancelled.")
                conn.close()
                return

            export_path = os.path.join(folder_selected, "DPR_Exported_Data_All.xlsx")

            data = cursor.fetchall()
            conn.close()

            processed_data = []
            for idx, row in enumerate(data, start=1):
                row = list(row)
                entered_by = row[-1]
                if entered_by and entered_by.strip().upper() == "ADMIN":
                    row[-2] = ""
                row = row[:-1]
                processed_data.append([idx] + row)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DPR Data"
            ws.append(headers)

            for row in processed_data:
                ws.append(row)

            try:
                wb.save(export_path)
                messagebox.showinfo("BMS DPR", f"Data exported successfully to:\n{export_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to save Excel file:\n{e}")
        else:
            if not self.selected_date:
                messagebox.showwarning("Export Error", "No date selected for export.")
                return

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Normalize casing and trim spaces for Entered_By check for Admin records
            try:
                cursor.execute("UPDATE DPR SET In_Time = '' WHERE TRIM(UPPER(Entered_By)) = 'ADMIN'")
                conn.commit()
            except Exception as e:
                messagebox.showerror("Database Error", f"Failed to update In Time for Admin records:\n{e}")
                conn.close()
                return

            # Query only rows for selected date
            try:
                cursor.execute("""
                    SELECT 
                        Lead, "Team Members", Project_Name, Process_Type, Other_process_details,
                        Attendance, Leave_Type, Half_Day_Status, Permission_Details, Count, Comments, Date, In_Time, Entered_By
                    FROM DPR
                    WHERE Date = ?
                """, (self.selected_date,))
            except Exception as e:
                messagebox.showerror("Database Error", f"Failed to fetch data for selected date:\n{e}")
                conn.close()
                return

            headers = [
                "SL.No", "Lead", "Team Members", "Project Name", "Process Type", "Other Process Details",
                "Attendance", "Leave Type", "Half Day Status", "Permission Details", "Count", "Comments", "Date", "In Time"
            ]

            folder_selected = filedialog.askdirectory(title="Select Folder to Save Excel File")
            if not folder_selected:
                messagebox.showwarning("Export Cancelled", "No folder was selected. Export cancelled.")
                conn.close()
                return

            export_date_str = datetime.strptime(self.selected_date, "%d-%m-%Y").strftime("%d%m%Y")
            export_path = os.path.join(folder_selected, f"DPR_Exported_Data_{export_date_str}.xlsx")

            data = cursor.fetchall()
            conn.close()

            processed_data = []
            for idx, row in enumerate(data, start=1):
                row = list(row)
                entered_by = row[-1]
                if entered_by and entered_by.strip().upper() == "ADMIN":
                    row[-2] = ""
                row = row[:-1]
                processed_data.append([idx] + row)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DPR Data"
            ws.append(headers)

            for row in processed_data:
                ws.append(row)

            try:
                wb.save(export_path)
                messagebox.showinfo("BMS DPR", f"Data exported successfully to:\n{export_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to save Excel file:\n{e}")



    def customize_export(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        self.customize_frame = ttk.Frame(self.root, padding=20)
        self.customize_frame.pack(expand=True, fill='both')

        ttk.Label(self.customize_frame, text="Select Custom Date Range", font=("Segoe UI", 20, "bold")).pack(pady=(10, 20))

        ttk.Label(self.customize_frame, text="From", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.from_date_entry_custom = DateEntry(self.customize_frame, width=30, bootstyle='info', firstweekday=6, dateformat="%d-%m-%Y")
        self.from_date_entry_custom.pack(pady=5)

        ttk.Label(self.customize_frame, text="To", font=("Segoe UI", 14)).pack(pady=(10, 0))
        self.to_date_entry_custom = DateEntry(self.customize_frame, width=30, bootstyle='info', firstweekday=6, dateformat="%d-%m-%Y")
        self.to_date_entry_custom.pack(pady=5)

        button_frame = ttk.Frame(self.customize_frame)
        button_frame.pack(pady=20)

        self.custom_back_button = ttk.Button(button_frame, text="← Back", command=self.back_to_export, style='BackForm.TButton')
        self.custom_back_button.pack(side='left', padx=10)

        self.custom_export_button = ttk.Button(button_frame, text="Export", command=self.export_customized_data, style='Export.TButton')
        self.custom_export_button.pack(side='left', padx=10)

    def back_to_export(self):
        self.customize_frame.destroy()
        self.create_widgets()

    def export_customized_data(self):
        from_date = self.from_date_entry_custom.entry.get()
        to_date = self.to_date_entry_custom.entry.get()

        if not from_date or not to_date:
            messagebox.showwarning("Input Error", "Please select both From and To dates.")
            return

        try:
            from_date_obj = datetime.strptime(from_date, "%d-%m-%Y")
            to_date_obj = datetime.strptime(to_date, "%d-%m-%Y")
        except Exception:
            messagebox.showwarning("Input Error", "Invalid date format selected.")
            return

        if from_date_obj > to_date_obj:
            messagebox.showwarning("Input Error", "From Date must be earlier than To Date.")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("UPDATE DPR SET In_Time = '' WHERE TRIM(UPPER(Entered_By)) = 'ADMIN'")
            conn.commit()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to update In Time for Admin records:\n{e}")
            conn.close()
            return

        from_date_sql = from_date_obj.strftime("%Y-%m-%d")
        to_date_sql = to_date_obj.strftime("%Y-%m-%d")

        date_convert_expr = "substr(Date,7,4) || '-' || substr(Date,4,2) || '-' || substr(Date,1,2)"

        try:
            cursor.execute(f"""
                SELECT 
                    Lead, "Team Members", Project_Name, Process_Type, Other_process_details,
                    Attendance, Leave_Type, Half_Day_Status, Permission_Details, Count, Comments, Date, In_Time, Entered_By
                FROM DPR
                WHERE {date_convert_expr} BETWEEN ? AND ?
            """, (from_date_sql, to_date_sql))
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to fetch data for selected date range:\n{e}")
            conn.close()
            return

        headers = [
            "SL.No", "Lead", "Team Members", "Project Name", "Process Type", "Other Process Details",
            "Attendance", "Leave Type", "Half Day Status", "Permission Details", "Count", "Comments", "Date", "In Time"
        ]

        folder_selected = filedialog.askdirectory(title="Select Folder to Save Excel File")
        if not folder_selected:
            messagebox.showwarning("Export Cancelled", "No folder was selected. Export cancelled.")
            conn.close()
            return

        from_date_file = from_date.replace('/', '_').replace('-', '_')
        to_date_file = to_date.replace('/', '_').replace('-', '_')
        export_path = os.path.join(folder_selected, f"DPR_Exported_Custom_Data_{from_date_file}_to_{to_date_file}.xlsx")

        data = cursor.fetchall()
        conn.close()

        processed_data = []
        for idx, row in enumerate(data, start=1):
            row = list(row)
            entered_by = row[-1]
            if entered_by and entered_by.strip().upper() == "ADMIN":
                row[-2] = ""
            row = row[:-1]
            processed_data.append([idx] + row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DPR Data"
        ws.append(headers)

        for row in processed_data:
            ws.append(row)

        try:
            wb.save(export_path)
            messagebox.showinfo("BMS DPR", f"Customized data exported successfully to:\n{export_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel file:\n{e}")

# ... [rest of your existing code]




    def export_monthly_data(self):
        dlg = MonthYearDialog(self.root)
        if dlg.result is None:
            return
        month, year = dlg.result

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Normalize casing and trim spaces for Entered_By check for Admin records if needed
        try:
            cursor.execute("UPDATE DPR SET In_Time = '' WHERE TRIM(UPPER(Entered_By)) = 'ADMIN'")
            conn.commit()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to update In Time for Admin records:\n{e}")
            conn.close()
            return

        # SQLite does not have direct substring on string, use substr function: substr(Date,4,2) gets month, substr(Date,7,4) gets year
        try:
            cursor.execute(f"""
                SELECT 
                    Lead, "Team Members", Project_Name, Process_Type, Other_process_details,
                    Attendance, Leave_Type, Half_Day_Status, Permission_Details, Count, Comments, Date, In_Time, Entered_By
                FROM DPR
                WHERE substr(Date,4,2) = ? AND substr(Date,7,4) = ?
            """, (month, year))
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to fetch data for selected month and year:\n{e}")
            conn.close()
            return

        data = cursor.fetchall()
        conn.close()

        if not data:
            messagebox.showinfo("No Data", f"No records found for {month}-{year}.")
            return

        headers = [
            "SL.No", "Lead", "Team Members", "Project Name", "Process Type", "Other Process Details",
            "Attendance", "Leave Type", "Half Day Status", "Permission Details", "Count", "Comments", "Date", "In Time"
        ]

        folder_selected = filedialog.askdirectory(title="Select Folder to Save Excel File")
        if not folder_selected:
            messagebox.showwarning("Export Cancelled", "No folder was selected. Export cancelled.")
            return

        export_path = os.path.join(folder_selected, f"DPR_Exported_Data_{month}{year}.xlsx")

        processed_data = []
        for idx, row in enumerate(data, start=1):
            row = list(row)
            entered_by = row[-1]
            if entered_by and entered_by.strip().upper() == "ADMIN":
                row[-2] = ""
            row = row[:-1]
            processed_data.append([idx] + row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DPR Data"
        ws.append(headers)

        for row in processed_data:
            ws.append(row)

        try:
            wb.save(export_path)
            messagebox.showinfo("BMS DPR", f"Monthly data exported successfully to:\n{export_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel file:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DPRApp(root)
    root.mainloop()